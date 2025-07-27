"""
Enhanced pytest configuration with additional fixtures for comprehensive Excel testing.
This extends your existing conftest.py with more test fixtures.
"""

import pytest
import os
from pathlib import Path
import tempfile
import openpyxl
from openpyxl import Workbook
from datetime import datetime, date


# Add these fixtures to your existing conftest.py file


@pytest.fixture(scope="session")
def comprehensive_xlsx_file(fixtures_dir):
    """Create a comprehensive test file with various data scenarios."""
    file_path = fixtures_dir / "comprehensive_test.xlsx"
    if not file_path.exists():
        create_comprehensive_xlsx(file_path)
    return str(file_path)


@pytest.fixture(scope="session")
def edge_cases_xlsx_file(fixtures_dir):
    """Create Excel file with edge case scenarios."""
    file_path = fixtures_dir / "edge_cases.xlsx"
    if not file_path.exists():
        create_edge_cases_xlsx(file_path)
    return str(file_path)


@pytest.fixture(scope="session")
def data_types_xlsx_file(fixtures_dir):
    """Create Excel file with various data types for testing."""
    file_path = fixtures_dir / "data_types.xlsx"
    if not file_path.exists():
        create_data_types_xlsx(file_path)
    return str(file_path)


@pytest.fixture(scope="session")
def corrupted_file_simulation(fixtures_dir):
    """Create a file that simulates corruption for error testing."""
    file_path = fixtures_dir / "corrupted.xlsx"
    if not file_path.exists():
        # Create a file with invalid Excel content
        with open(file_path, 'w') as f:
            f.write("This is definitely not a valid Excel file content")
    return str(file_path)


@pytest.fixture
def temp_excel_file():
    """Create a temporary Excel file that gets cleaned up after test."""
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        temp_path = tmp.name
    
    # Create a basic Excel file
    wb = Workbook()
    ws = wb.active
    ws.title = "TempSheet"
    ws.append(["Temp", "Data"])
    ws.append(["Test", "Value"])
    wb.save(temp_path)
    
    yield temp_path
    
    # Cleanup
    try:
        os.unlink(temp_path)
    except (OSError, FileNotFoundError):
        pass  # File might already be deleted


# Helper functions to create specialized test files

def create_comprehensive_xlsx(file_path):
    """Create a comprehensive Excel file with multiple scenarios."""
    file_path.parent.mkdir(parents=True, exist_ok=True)
    
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Sheet 1: Standard data
    ws1 = wb.create_sheet("StandardData")
    ws1.append(["ID", "Name", "Department", "Salary", "Start Date"])
    ws1.append([1, "Alice Johnson", "Engineering", 75000, date(2020, 1, 15)])
    ws1.append([2, "Bob Smith", "Marketing", 65000, date(2019, 3, 22)])
    ws1.append([3, "Carol Williams", "HR", 55000, date(2021, 7, 10)])
    ws1.append([4, "David Brown", "Finance", 70000, date(2018, 11, 5)])
    
    # Sheet 2: Data with empty rows/columns
    ws2 = wb.create_sheet("WithEmptyRows")
    ws2.append(["Column1", "Column2", "Column3"])
    ws2.append(["Data1", "Data2", "Data3"])
    # Skip row 3 (empty)
    ws2.append([None, None, None])  # Row 4: all empty
    ws2.append(["Data4", "", "Data6"])  # Row 5: partial data
    ws2.append(["", "", ""])  # Row 6: empty strings
    ws2.append(["Final", "Row", "Data"])  # Row 7: valid data
    
    # Sheet 3: Numeric data with calculations
    ws3 = wb.create_sheet("NumericData")
    ws3.append(["Value1", "Value2", "Sum", "Product"])
    ws3.append([10, 20, 30, 200])
    ws3.append([5, 15, 20, 75])
    ws3.append([100, 50, 150, 5000])
    
    # Sheet 4: Mixed data types
    ws4 = wb.create_sheet("MixedTypes")
    ws4.append(["String", "Integer", "Float", "Boolean", "Date"])
    ws4.append(["Text", 42, 3.14, True, date(2023, 1, 1)])
    ws4.append(["Another", 0, -2.5, False, date(2023, 12, 31)])
    
    wb.save(file_path)


def create_edge_cases_xlsx(file_path):
    """Create Excel file with edge case scenarios."""
    file_path.parent.mkdir(parents=True, exist_ok=True)
    
    wb = Workbook()
    wb.remove(wb.active)
    
    # Sheet 1: Very long text
    ws1 = wb.create_sheet("LongText")
    long_text = "This is a very long text that goes on and on " * 50
    ws1.append(["ID", "Long Text"])
    ws1.append([1, long_text])
    ws1.append([2, "Short text"])
    
    # Sheet 2: Special characters
    ws2 = wb.create_sheet("SpecialChars")
    ws2.append(["Type", "Content"])
    ws2.append(["Quotes", 'Text with "quotes" and \'apostrophes\''])
    ws2.append(["Commas", "Text, with, commas, everywhere"])
    ws2.append(["Newlines", "Text\nwith\nnewlines"])
    ws2.append(["Unicode", "Ñoñó 中文 العربية 🎉"])
    ws2.append(["Special", "!@#$%^&*()_+-=[]{}|;:,.<>?"])
    
    # Sheet 3: Numbers as text and edge numeric values
    ws3 = wb.create_sheet("NumericEdges")
    ws3.append(["Description", "Value"])
    ws3['A2'] = "Very large number"
    ws3['B2'] = 999999999999999999
    ws3['A3'] = "Very small decimal"
    ws3['B3'] = 0.000000000001
    ws3['A4'] = "Zero"
    ws3['B4'] = 0
    ws3['A5'] = "Negative"
    ws3['B5'] = -12345
    
    # Sheet 4: Empty sheet (only title)
    ws4 = wb.create_sheet("EmptyContent")
    # Don't add any content
    
    wb.save(file_path)


def create_data_types_xlsx(file_path):
    """Create Excel file focused on testing different data types."""
    file_path.parent.mkdir(parents=True, exist_ok=True)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "DataTypes"
    
    # Header
    ws.append(["Type", "Value", "Description"])
    
    # Different data types
    ws.append(["String", "Hello World", "Basic text"])
    ws.append(["Integer", 42, "Whole number"])
    ws.append(["Float", 3.14159, "Decimal number"])
    ws.append(["Boolean", True, "Boolean true"])
    ws.append(["Boolean", False, "Boolean false"])
    ws.append(["Date", date(2023, 6, 15), "Date value"])
    ws.append(["DateTime", datetime(2023, 6, 15, 14, 30, 0), "DateTime value"])
    ws.append(["Null", None, "Null value"])
    ws.append(["Empty String", "", "Empty string"])
    ws.append(["Whitespace", "   ", "Only whitespace"])
    ws.append(["Zero", 0, "Numeric zero"])
    ws.append(["Large Number", 1234567890123456789, "Very large number"])
    
    wb.save(file_path)


def create_performance_test_xlsx(file_path, rows=1000):
    """Create a large Excel file for performance testing."""
    file_path.parent.mkdir(parents=True, exist_ok=True)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "PerformanceTest"
    
    # Add header
    ws.append(["ID", "Name", "Category", "Value", "Timestamp"])
    
    # Add many rows of data
    import random
    categories = ["Alpha", "Beta", "Gamma", "Delta", "Epsilon"]
    
    for i in range(1, rows + 1):
        row = [
            i,
            f"Item_{i:06d}",
            random.choice(categories),
            random.randint(1, 10000),
            datetime(2023, 1, 1) + (datetime(2023, 12, 31) - datetime(2023, 1, 1)) * random.random()
        ]
        ws.append(row)
        
        # Progress indicator for large files
        if i % 100 == 0:
            print(f"Created {i}/{rows} rows")
    
    wb.save(file_path)


# Additional fixture for performance testing
@pytest.fixture(scope="session")
def performance_xlsx_file(fixtures_dir):
    """Create a large Excel file for performance testing."""
    file_path = fixtures_dir / "performance_test.xlsx"
    if not file_path.exists():
        print("Creating performance test file (this may take a moment)...")
        create_performance_test_xlsx(file_path, rows=2000)
    return str(file_path)


# Fixture for testing file permissions (Unix-like systems)
@pytest.fixture
def readonly_excel_file(temp_excel_file):
    """Create a read-only Excel file for permission testing."""
    # Make file read-only
    import stat
    os.chmod(temp_excel_file, stat.S_IREAD)
    
    yield temp_excel_file
    
    # Restore permissions for cleanup
    try:
        os.chmod(temp_excel_file, stat.S_IWRITE | stat.S_IREAD)
    except (OSError, FileNotFoundError):
        pass


# Test data samples for validation
@pytest.fixture
def expected_sample_data():
    """Provide expected data structure for validation tests."""
    return {
        'employees_count': 3,
        'summary_metrics': 4,
        'expected_names': ['John Doe', 'Jane Smith', 'Bob Johnson'],
        'expected_departments': ['Engineering', 'Design', 'Management'],
        'expected_worksheets': ['Employees', 'Summary', 'Mixed Data']
    }


# Utility fixture for creating custom test files on the fly
@pytest.fixture
def excel_file_factory(tmp_path):
    """Factory for creating custom Excel files during tests."""
    def _create_file(filename, sheets_data):
        """
        Create Excel file with custom data.
        
        Args:
            filename: Name of the file
            sheets_data: Dict with sheet_name -> list_of_rows
        """
        file_path = tmp_path / filename
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        for sheet_name, rows in sheets_data.items():
            ws = wb.create_sheet(sheet_name)
            for row in rows:
                ws.append(row)
        
        wb.save(file_path)
        return str(file_path)
    
    return _create_file