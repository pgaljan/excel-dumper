"""
Fixed pytest configuration with all required fixtures for excel_dumper tests.
Replace your conftest.py with this content.
"""

import pytest
import os
from pathlib import Path
import tempfile
import openpyxl
from openpyxl import Workbook
from datetime import datetime, date


# Test data directory paths
TEST_DIR = Path(__file__).parent
FIXTURES_DIR = TEST_DIR / "fixtures"


@pytest.fixture(scope="session")
def fixtures_dir():
    """Provide path to test fixtures directory."""
    FIXTURES_DIR.mkdir(parents=True, exist_ok=True)
    return FIXTURES_DIR


@pytest.fixture(scope="session")
def sample_xlsx_file(fixtures_dir):
    """Provide path to sample XLSX file."""
    file_path = fixtures_dir / "sample.xlsx"
    if not file_path.exists():
        create_sample_xlsx(file_path)
    return str(file_path)


@pytest.fixture(scope="session")
def formulas_xlsx_file(fixtures_dir):
    """Provide path to Excel file with formulas."""
    file_path = fixtures_dir / "formulas.xlsx"
    if not file_path.exists():
        create_formulas_xlsx(file_path)
    return str(file_path)


@pytest.fixture(scope="session")
def hidden_sheets_file(fixtures_dir):
    """Provide path to Excel file with hidden sheets."""
    file_path = fixtures_dir / "hidden_sheets.xlsx"
    if not file_path.exists():
        create_hidden_sheets_xlsx(file_path)
    return str(file_path)


@pytest.fixture(scope="session")
def empty_xlsx_file(fixtures_dir):
    """Provide path to empty Excel file."""
    file_path = fixtures_dir / "empty.xlsx"
    if not file_path.exists():
        create_empty_xlsx(file_path)
    return str(file_path)


@pytest.fixture(scope="session")
def large_xlsx_file(fixtures_dir):
    """Provide path to large Excel file."""
    file_path = fixtures_dir / "large_file.xlsx"
    if not file_path.exists():
        create_large_xlsx(file_path)
    return str(file_path)


@pytest.fixture
def temp_excel_file():
    """Create a temporary Excel file that gets cleaned up after test."""
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        temp_path = tmp.name
    
    # Create a basic Excel file
    wb = Workbook()
    ws = wb.active
    ws.title = "TempSheet"
    ws.append(["Temp", "Data"])
    ws.append(["Test", "Value"])
    wb.save(temp_path)
    
    yield temp_path
    
    # Cleanup
    try:
        os.unlink(temp_path)
    except (OSError, FileNotFoundError):
        pass  # File might already be deleted


@pytest.fixture
def excel_file_factory(tmp_path):
    """Factory for creating custom Excel files during tests."""
    def _create_file(filename, sheets_data):
        """
        Create Excel file with custom data.
        
        Args:
            filename: Name of the file
            sheets_data: Dict with sheet_name -> list_of_rows
        """
        file_path = tmp_path / filename
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        for sheet_name, rows in sheets_data.items():
            ws = wb.create_sheet(sheet_name)
            for row in rows:
                ws.append(row)
        
        wb.save(file_path)
        return str(file_path)
    
    return _create_file


# Helper functions to create test Excel files

def create_sample_xlsx(file_path):
    """Create a sample XLSX file with test data."""
    file_path.parent.mkdir(parents=True, exist_ok=True)
    
    wb = Workbook()
    
    # Remove default sheet and create our sheets
    wb.remove(wb.active)
    
    # Sheet1: Employees
    ws1 = wb.create_sheet("Employees")
    ws1.append(["Name", "Department", "Age", "Salary"])
    ws1.append(["John Doe", "Engineering", 30, 75000])
    ws1.append(["Jane Smith", "Design", 25, 65000])
    ws1.append(["Bob Johnson", "Management", 35, 85000])
    
    # Sheet2: Summary
    ws2 = wb.create_sheet("Summary")
    ws2.append(["Metric", "Value"])
    ws2.append(["Total Employees", 3])
    ws2.append(["Average Age", 30])
    ws2.append(["Average Salary", 75000])
    
    # Sheet3: Mixed Data
    ws3 = wb.create_sheet("Mixed Data")
    ws3.append(["Text", "Number", "Date", "Boolean", "Empty"])
    ws3.append(["Sample", 42, "2023-01-01", True, None])
    ws3.append(["Test", 0, "2023-12-31", False, ""])
    ws3.append(["", 999, None, None, "Value"])
    
    wb.save(file_path)


def create_formulas_xlsx(file_path):
    """Create XLSX file with formulas for testing formula extraction."""
    file_path.parent.mkdir(parents=True, exist_ok=True)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Formulas"
    
    # Add some data
    ws['A1'] = "Numbers"
    ws['A2'] = 10
    ws['A3'] = 20
    ws['A4'] = 30
    
    # Add formulas
    ws['B1'] = "Calculations"
    ws['B2'] = "=A2*2"
    ws['B3'] = "=SUM(A2:A4)"
    ws['B4'] = "=AVERAGE(A2:A4)"
    ws['B5'] = "=IF(A2>15,\"High\",\"Low\")"
    
    wb.save(file_path)


def create_hidden_sheets_xlsx(file_path):
    """Create XLSX file with hidden sheets."""
    file_path.parent.mkdir(parents=True, exist_ok=True)
    
    wb = Workbook()
    
    # Visible sheet
    ws1 = wb.active
    ws1.title = "Visible"
    ws1.append(["Visible Data", "Column 2"])
    ws1.append(["Row 1", "Data 1"])
    ws1.append(["Row 2", "Data 2"])
    
    # Hidden sheet
    ws2 = wb.create_sheet("Hidden")
    ws2.append(["Hidden Data", "Secret Info"])
    ws2.append(["Secret 1", "Confidential"])
    ws2.append(["Secret 2", "Internal"])
    ws2.sheet_state = 'hidden'
    
    wb.save(file_path)


def create_empty_xlsx(file_path):
    """Create an empty XLSX file."""
    file_path.parent.mkdir(parents=True, exist_ok=True)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Empty"
    # Don't add any data - leave completely empty
    
    wb.save(file_path)


def create_large_xlsx(file_path):
    """Create a large XLSX file for performance testing."""
    file_path.parent.mkdir(parents=True, exist_ok=True)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "LargeData"
    
    # Add header
    ws.append(["ID", "Name", "Value", "Category"])
    
    # Add 1000 rows of data
    import random
    
    categories = ["A", "B", "C", "D", "E"]
    
    for i in range(1, 1001):
        row = [
            i,
            f"Item_{i:04d}",
            random.randint(1, 1000),
            random.choice(categories)
        ]
        ws.append(row)
    
    wb.save(file_path)


def pytest_configure(config):
    """Ensure fixtures directory exists before tests run."""
    FIXTURES_DIR.mkdir(parents=True, exist_ok=True)