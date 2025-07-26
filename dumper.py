#!/usr/bin/env python3
"""
Excel Sheet Dumper Script
Extracts all non-null rows from Excel worksheets and saves to CSV or JSON format.
"""

import argparse
import csv
import glob
import json
import os
import sys
from datetime import datetime
from pathlib import Path

try:
    import pandas as pd
    import openpyxl
except ImportError as e:
    print("Error: Required libraries not found.")
    print("Please install with: pip install pandas openpyxl xlrd")
    print(f"Missing: {e}")
    sys.exit(1)


def find_newest_excel_file(input_dir="."):
    """Find the newest Excel file in the specified directory."""
    excel_patterns = ['*.xlsx', '*.xls', '*.xlsm', '*.xlsb']
    excel_files = []
    
    # Change to the specified directory for globbing
    search_path = Path(input_dir)
    
    for pattern in excel_patterns:
        excel_files.extend(search_path.glob(pattern))
    
    if not excel_files:
        raise FileNotFoundError(f"No Excel files found in directory: {input_dir}")
    
    # Get the newest file based on modification time
    newest_file = max(excel_files, key=lambda f: f.stat().st_mtime)
    return str(newest_file)


def has_non_null_data(row):
    """Check if a row contains any non-null data."""
    return any(cell is not None and str(cell).strip() != '' for cell in row)


def extract_excel_data(filename, include_hidden=True, include_row_numbers=False, include_formulas=False):
    """Extract data from all worksheets in an Excel file."""
    try:
        extracted_data = []
        
        # Check for unsupported file types with formulas option
        if include_formulas and not filename.lower().endswith(('.xlsx', '.xlsm')):
            print(f"Warning: -formulas option only works with .xlsx and .xlsm files.")
            print(f"File '{filename}' will be processed with calculated values instead of formulas.")
            include_formulas = False  # Disable formulas for unsupported files
        
        # If formulas are requested, we need to use openpyxl for direct cell access
        if include_formulas and filename.lower().endswith(('.xlsx', '.xlsm')):
            wb = openpyxl.load_workbook(filename, data_only=False)
            
            for sheet_name in wb.sheetnames:
                try:
                    sheet = wb[sheet_name]
                    
                    # Check if sheet is hidden
                    if not include_hidden and sheet.sheet_state == 'hidden':
                        print(f"Skipping hidden sheet: {sheet_name}")
                        continue
                    
                    # Process each row in the sheet
                    for row_idx, row in enumerate(sheet.iter_rows(values_only=False), 1):
                        row_data = []
                        has_data = False
                        
                        for cell in row:
                            if cell.value is not None:
                                # Check if cell has a formula
                                if hasattr(cell, 'formula') and cell.formula:
                                    # Prefix with "FORMULA: " to prevent CSV interpretation and circular references
                                    row_data.append(f"FORMULA: ={cell.formula}")
                                else:
                                    row_data.append(cell.value)
                                has_data = True
                            else:
                                row_data.append(None)
                        
                        if has_data and has_non_null_data(row_data):
                            # Build the output row
                            if include_row_numbers:
                                row_with_metadata = [sheet_name, row_idx] + row_data
                            else:
                                row_with_metadata = [sheet_name] + row_data
                            
                            extracted_data.append(row_with_metadata)
                            
                except Exception as e:
                    print(f"Warning: Could not process sheet '{sheet_name}': {e}")
                    continue
        
        else:
            # Use pandas for standard data extraction (calculated values)
            excel_file = pd.ExcelFile(filename)
            
            for sheet_name in excel_file.sheet_names:
                try:
                    # Read the sheet into a DataFrame
                    df = pd.read_excel(filename, sheet_name=sheet_name, header=None)
                    
                    # Skip empty sheets
                    if df.empty:
                        continue
                    
                    # Check if sheet is hidden (requires openpyxl for .xlsx files)
                    if not include_hidden:
                        try:
                            # Load workbook to check sheet visibility
                            if filename.lower().endswith(('.xlsx', '.xlsm')):
                                wb = openpyxl.load_workbook(filename)
                                if sheet_name in wb.sheetnames:
                                    sheet = wb[sheet_name]
                                    if sheet.sheet_state == 'hidden':
                                        print(f"Skipping hidden sheet: {sheet_name}")
                                        continue
                        except Exception:
                            # If we can't check visibility, include the sheet
                            pass
                    
                    # Convert DataFrame to list of lists and process each row
                    for row_idx, row in df.iterrows():
                        row_data = row.tolist()
                        if has_non_null_data(row_data):
                            # Build the output row
                            if include_row_numbers:
                                # Excel rows are 1-indexed, and we add 1 to account for pandas 0-indexing
                                excel_row_number = row_idx + 1
                                row_with_metadata = [sheet_name, excel_row_number] + row_data
                            else:
                                row_with_metadata = [sheet_name] + row_data
                            
                            extracted_data.append(row_with_metadata)
                            
                except Exception as e:
                    print(f"Warning: Could not process sheet '{sheet_name}': {e}")
                    continue
        
        return extracted_data
        
    except Exception as e:
        raise Exception(f"Error reading Excel file '{filename}': {str(e)}")


def write_to_csv(data, output_filename, include_row_numbers=False):
    """Write extracted data to CSV file."""
    try:
        with open(output_filename, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            
            # Write header
            if data:
                # Determine maximum number of columns in the data
                max_cols = max(len(row) for row in data) if data else 0
                
                if include_row_numbers:
                    # Structure: [worksheet_name, row_number, ...data_columns...]
                    # So data columns = total columns - 2
                    data_cols = max_cols - 2
                    if data_cols > 0:
                        header = ['Worksheet', 'Row_Number'] + [f'Column_{i}' for i in range(1, data_cols + 1)]
                    else:
                        header = ['Worksheet', 'Row_Number']
                else:
                    # Structure: [worksheet_name, ...data_columns...]
                    # So data columns = total columns - 1
                    data_cols = max_cols - 1
                    if data_cols > 0:
                        header = ['Worksheet'] + [f'Column_{i}' for i in range(1, data_cols + 1)]
                    else:
                        header = ['Worksheet']
                
                writer.writerow(header)
            
            # Write data rows
            for row in data:
                writer.writerow(row)
                
        print(f"Data successfully exported to: {output_filename}")
        print(f"Total rows exported: {len(data)}")
        
    except Exception as e:
        raise Exception(f"Error writing to CSV file '{output_filename}': {str(e)}")


def write_to_json(data, output_filename, include_row_numbers=False):
    """Write extracted data to JSON file, excluding null values."""
    try:
        json_data = []
        
        if data:
            # Determine maximum number of columns in the data
            max_cols = max(len(row) for row in data) if data else 0
            
            # Create column names based on structure
            if include_row_numbers:
                # Structure: [worksheet_name, row_number, ...data_columns...]
                data_cols = max_cols - 2
                if data_cols > 0:
                    column_names = ['Worksheet', 'Row_Number'] + [f'Column_{i}' for i in range(1, data_cols + 1)]
                else:
                    column_names = ['Worksheet', 'Row_Number']
            else:
                # Structure: [worksheet_name, ...data_columns...]
                data_cols = max_cols - 1
                if data_cols > 0:
                    column_names = ['Worksheet'] + [f'Column_{i}' for i in range(1, data_cols + 1)]
                else:
                    column_names = ['Worksheet']
            
            # Convert each row to a dictionary, excluding null values
            for row in data:
                row_dict = {}
                for i, value in enumerate(row[:len(column_names)]):
                    # Skip null values, empty strings, and pandas NA values
                    if (value is not None and 
                        not (hasattr(value, 'isna') and value.isna()) and 
                        str(value).lower() != 'nan' and 
                        str(value).strip() != ''):
                        row_dict[column_names[i]] = value
                json_data.append(row_dict)
        
        # Write JSON with pretty formatting
        with open(output_filename, 'w', encoding='utf-8') as jsonfile:
            json.dump(json_data, jsonfile, indent=2, ensure_ascii=False, default=str)
                
        print(f"Data successfully exported to: {output_filename}")
        print(f"Total rows exported: {len(data)}")
        
    except Exception as e:
        raise Exception(f"Error writing to JSON file '{output_filename}': {str(e)}")


def generate_output_filename(input_filename, output_dir=None, output_format='csv'):
    """Generate output filename based on input filename with timestamp."""
    input_path = Path(input_filename)
    base_name = input_path.stem
    
    # Get file modification time with timezone
    mod_time = os.path.getmtime(input_filename)
    mod_datetime = datetime.fromtimestamp(mod_time).astimezone()
    
    # Format timestamp as ISO 8601 (e.g., 2025-07-21T14:30:52-05:00)
    # Replace colons with hyphens for filename compatibility
    timestamp = mod_datetime.isoformat().replace(':', '-')
    
    base_filename = f"dumperpy_{base_name}_{timestamp}"
    
    # Determine the directory path
    if output_dir:
        output_path = Path(output_dir)
        # Create directory if it doesn't exist
        output_path.mkdir(parents=True, exist_ok=True)
        base_path = output_path / base_filename
    else:
        base_path = Path(base_filename)
    
    # Set file extension based on format
    file_extension = '.json' if output_format == 'json' else '.csv'
    
    # Check if file exists and find an available filename
    counter = 0
    while True:
        if counter == 0:
            final_filename = f"{base_path}{file_extension}"
        else:
            final_filename = f"{base_path}({counter}){file_extension}"
        
        if not Path(final_filename).exists():
            return str(final_filename)
        
        counter += 1


def show_help():
    """Display help information."""
    help_text = """
Excel Sheet Dumper - Extract data from Excel worksheets to CSV or JSON

USAGE:
    python dumper.py [OPTIONS]

OPTIONS:
    -file FILE          Specify Excel file to process (default: newest Excel file in input directory)
    -input DIR         Input directory to search for Excel files (default: current directory)
    -output DIR        Output directory for output file (default: current directory)
    -no-hide           Skip hidden worksheets (default: include all worksheets)
    -rownumbers        Include Excel row numbers in output (default: exclude row numbers)
    -formulas          Show formulas instead of calculated values (.xlsx/.xlsm only)
    -json              Output to JSON format instead of CSV (default: CSV)
    -help              Show this help message

EXAMPLES:
    python dumper.py                    # Process newest Excel file from current directory to CSV
    python dumper.py -file data.xlsx    # Process specific file to CSV
    python dumper.py -json              # Process newest Excel file to JSON format
    python dumper.py -file data.xlsx -json  # Process specific file to JSON format
    python dumper.py -rownumbers        # Include Excel row numbers in CSV output
    python dumper.py -formulas          # Show formulas instead of values in CSV
    python dumper.py -input ./source    # Process newest file from ./source directory
    python dumper.py -input ./source -output ./exports  # Source and output directories
    python dumper.py -input /data -file report.xlsx -rownumbers -json  # Specific file with row numbers to JSON
    python dumper.py -file data.xlsx -output ./exports -no-hide -rownumbers -formulas -json  # All options combined

OUTPUT:
    Creates a CSV or JSON file named "dumperpy_[original_filename]_[timestamp].[csv|json]" with:
    - Timestamp is the last modified time of the originating Excel file
    - Timestamp format: ISO 8601 with colons replaced by hyphens (e.g., dumperpy_data_2025-07-21T14-30-52-05-00.csv)
    - If file exists, appends incremental number in parentheses (e.g., dumperpy_data_2025-07-21T14-30-52-05-00(1).csv)
    
    CSV FORMAT:
    - First column: Worksheet name
    - Second column: Excel row number (if -rownumbers option used)
    - Remaining columns: Original data from worksheets
    - Only non-empty rows are included
    
    JSON FORMAT:
    - Array of objects where each object represents a row
    - Each object has keys: "Worksheet", "Row_Number" (if -rownumbers used), "Column_1", "Column_2", etc.
    - Only non-empty rows are included
    - Pretty-formatted with 2-space indentation
    - Null values and empty strings are excluded from JSON objects
    
    - Formulas are prefixed with 'FORMULA: =' to prevent circular references in spreadsheet applications

NOTE: The -formulas option only works with .xlsx and .xlsm files. For .xls and .xlsb files, 
calculated values will be shown regardless of this setting.

PYTHON DEPENDENCIES:
    - pandas         (pip install pandas)
    - openpyxl       (pip install openpyxl) - for .xlsx/.xlsm files
    - xlrd           (pip install xlrd) - for .xls files
    - Standard library: argparse, csv, json, glob, os, sys, pathlib, datetime

    Install all at once: pip install pandas openpyxl xlrd

SUPPORTED EXCEL FORMATS:
    - .xlsx (Excel 2007+)
    - .xls  (Excel 97-2003)
    - .xlsm (Excel Macro-Enabled)
    - .xlsb (Excel Binary)
"""
    print(help_text)


def main():
    """Main function to handle command line arguments and orchestrate the process."""
    
    # Handle -help parameter separately since argparse's help conflicts with our custom help
    if '-help' in sys.argv:
        show_help()
        return
    
    parser = argparse.ArgumentParser(description='Extract Excel worksheet data to CSV or JSON', add_help=False)
    parser.add_argument('-file', dest='filename', help='Excel file to process')
    parser.add_argument('-no-hide', action='store_true', help='Skip hidden worksheets')
    parser.add_argument('-output', dest='output_dir', help='Output directory for output file')
    parser.add_argument('-input', dest='input_dir', help='Input directory to search for Excel files')
    parser.add_argument('-formulas', action='store_true', help='Show formulas instead of calculated values (.xlsx/.xlsm only)')
    parser.add_argument('-rownumbers', action='store_true', help='Include Excel row numbers in output')
    parser.add_argument('-json', action='store_true', help='Output to JSON format instead of CSV')
    
    try:
        args = parser.parse_args()
        
        # Determine output format
        output_format = 'json' if args.json else 'csv'
        
        # Determine input file
        if args.filename:
            # If filename is provided, check if it's absolute or relative
            file_path = Path(args.filename)
            if not file_path.is_absolute() and args.input_dir:
                # If it's relative and input_dir is specified, join them
                input_file = str(Path(args.input_dir) / args.filename)
            else:
                input_file = args.filename
                
            if not os.path.exists(input_file):
                print(f"Error: File '{input_file}' not found.")
                sys.exit(1)
        else:
            try:
                input_dir = args.input_dir if args.input_dir else "."
                input_file = find_newest_excel_file(input_dir)
                print(f"Processing newest Excel file: {Path(input_file).name}")
                print(f"From directory: {Path(input_file).parent}")
            except FileNotFoundError as e:
                print(f"Error: {e}")
                sys.exit(1)
        
        # Extract data
        include_hidden = not args.no_hide
        include_row_numbers = args.rownumbers
        include_formulas = args.formulas
        print(f"Extracting data from: {input_file}")
        print(f"Including hidden sheets: {include_hidden}")
        print(f"Including row numbers: {include_row_numbers}")
        print(f"Including formulas: {include_formulas}")
        print(f"Output format: {output_format.upper()}")
        
        extracted_data = extract_excel_data(input_file, include_hidden, include_row_numbers, include_formulas)
        
        if not extracted_data:
            print("No data found to export.")
            return
        
        # Generate output filename and write file
        output_file = generate_output_filename(input_file, args.output_dir, output_format)
        
        if output_format == 'json':
            write_to_json(extracted_data, output_file, include_row_numbers)
        else:
            write_to_csv(extracted_data, output_file, include_row_numbers)
        
    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()